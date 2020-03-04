import openpyxl

wb = openpyxl.load_workbook( r'H:\Python\Practice\Output_File.xlsx' )
sheets = wb.sheetnames
print( f"len( sheets ) = { len( sheets ) }" )
print( sheets )

for i in sheets:
    sheet = wb[ i ]
    print( sheet )
    m_row=sheet.max_row
    m_col=sheet.max_column
    
    print( f"Sheet number = { sheet.title }" )
    sheet_rows_columns = list( sheet.columns )

    
    for sheet_column in sheet_rows_columns:
        for block in sheet_column:
            
            block = str(block)
            start_index = block.index('.')
            end_index = block.index('>')
            
            block_name = block[ start_index + 1 : end_index ]
            print( sheet[block_name].value, end = ', ' )
        print()

    print()
